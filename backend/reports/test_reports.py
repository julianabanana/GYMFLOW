"""
Tests de 010-reportes-asistencia contra los criterios de aceptación de
spec/features/010-reportes-asistencia/spec.md: filtro por rango, unidad de
conteo (día de asistencia, no reingresos), export XLSX/CSV con los mismos
datos, RBAC solo Administrador, inmutabilidad y rango vacío/ inválido.
"""
import csv
import io
from datetime import timedelta

from fastapi.testclient import TestClient

from auth.conftest import _crear_staff
from core.config import now as _now
from core.security import create_access_token
from main import app
from membership.service import hoy
from reports.schemas import ReportFilters
from reports.service import export, generate
from models import (
    CheckIn,
    EstadoMembresia,
    EstadoUsuario,
    Membership,
    MembershipType,
    ResultadoCheckin,
    RolUsuario,
    User,
)

client = TestClient(app)


def _token(user) -> dict:
    return {"Authorization": f"Bearer {create_access_token({'sub': str(user.id), 'rol': user.rol.value})}"}


def _crear_socio(db, cedula, nombre, tipo_nombre="Mensual"):
    tipo = MembershipType(
        nombre=tipo_nombre, precio_base=50000, visitas_totales=30,
        cupo_invitados=1, duracion_dias=30, activo=True,
    )
    db.add(tipo)
    db.flush()
    user = User(cedula=cedula, nombre=nombre, rol=RolUsuario.miembro, estado=EstadoUsuario.activo)
    db.add(user)
    db.flush()
    db.add(Membership(
        miembro_id=user.id, tipo_id=tipo.id, visitas_restantes=10,
        cupo_invitados_restantes=1, fecha_inicio=hoy(),
        fecha_vencimiento=hoy() + timedelta(days=30),
        estado=EstadoMembresia.activa, monto=tipo.precio_base,
    ))
    db.flush()
    return user


def _checkin(db, user, cuando, is_active=True, resultado=ResultadoCheckin.exitoso):
    c = CheckIn(usuario_id=user.id, fecha_hora=cuando, resultado=resultado, is_active=is_active)
    db.add(c)
    db.flush()
    return c


def test_generate_filtra_por_rango(db):
    user = _crear_socio(db, "3000000001", "Ana")
    _checkin(db, user, _now() - timedelta(days=10))  # fuera del rango
    _checkin(db, user, _now() - timedelta(days=2))   # dentro
    db.commit()

    filtros = ReportFilters(fecha_inicio=hoy() - timedelta(days=5), fecha_fin=hoy())
    filas = generate(filtros, db)

    assert len(filas) == 1
    assert filas[0].usuario_nombre == "Ana"
    assert filas[0].tipo_membresia == "Mensual"
    assert filas[0].resultado == "exitoso"


def test_generate_rango_vacio_devuelve_lista(db):
    filtros = ReportFilters(fecha_inicio=hoy() - timedelta(days=5), fecha_fin=hoy())
    assert generate(filtros, db) == []


def test_generate_reingreso_mismo_dia_no_cuenta(db):
    user = _crear_socio(db, "3000000002", "Beto")
    _checkin(db, user, _now(), is_active=True)
    # reingreso el mismo día: is_active=false → no suma como asistencia
    _checkin(db, user, _now() + timedelta(minutes=5), is_active=False)
    db.commit()

    filtros = ReportFilters(fecha_inicio=hoy(), fecha_fin=hoy())
    filas = generate(filtros, db)

    assert len(filas) == 1


def test_generate_no_incluye_denegados(db):
    user = _crear_socio(db, "3000000003", "Caro")
    _checkin(db, user, _now(), is_active=False, resultado=ResultadoCheckin.denegado)
    db.commit()

    filtros = ReportFilters(fecha_inicio=hoy(), fecha_fin=hoy())
    assert generate(filtros, db) == []


def test_generate_no_muta_checkins(db):
    user = _crear_socio(db, "3000000004", "Dani")
    _checkin(db, user, _now())
    db.commit()
    antes = db.query(CheckIn).count()

    filtros = ReportFilters(fecha_inicio=hoy(), fecha_fin=hoy())
    generate(filtros, db)

    assert db.query(CheckIn).count() == antes


def test_export_csv_y_xlsx_mismos_datos(db):
    user = _crear_socio(db, "3000000005", "Eva")
    _checkin(db, user, _now())
    db.commit()
    filtros = ReportFilters(fecha_inicio=hoy(), fecha_fin=hoy())

    csv_bytes = export(filtros, "csv", db)
    xlsx_bytes = export(filtros, "xlsx", db)

    reader = list(csv.reader(io.StringIO(csv_bytes.decode("utf-8-sig"))))
    assert reader[0][0] == "Fecha y hora"
    assert reader[1][1] == "Eva"
    assert len(reader) == 2  # encabezado + 1 asistencia

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    valores = list(ws.iter_rows(values_only=True))
    assert valores[0][0] == "Fecha y hora"
    assert valores[1][1] == "Eva"
    assert len(valores) == 2


def test_router_attendance_solo_admin(db):
    empleado = _crear_staff(db, RolUsuario.empleado, "emp-rep@gymflow.test")
    params = {"fecha_inicio": str(hoy()), "fecha_fin": str(hoy())}

    assert client.get("/reportes/attendance", params=params).status_code == 401
    assert client.get("/reportes/attendance", params=params, headers=_token(empleado)).status_code == 403


def test_router_attendance_admin_ok(db):
    user = _crear_socio(db, "3000000006", "Fabi")
    _checkin(db, user, _now())
    db.commit()
    admin = _crear_staff(db, RolUsuario.administrador, "admin-rep@gymflow.test")
    params = {"fecha_inicio": str(hoy()), "fecha_fin": str(hoy())}

    resp = client.get("/reportes/attendance", params=params, headers=_token(admin))

    assert resp.status_code == 200
    body = resp.json()
    assert len(body) == 1
    assert body[0]["usuario_nombre"] == "Fabi"


def test_router_export_xlsx_headers(db):
    admin = _crear_staff(db, RolUsuario.administrador, "admin-exp@gymflow.test")
    params = {"fecha_inicio": str(hoy()), "fecha_fin": str(hoy()), "format": "xlsx"}

    resp = client.get("/reportes/attendance/export", params=params, headers=_token(admin))

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]


def test_router_rango_invalido_422(db):
    admin = _crear_staff(db, RolUsuario.administrador, "admin-inv@gymflow.test")
    params = {"fecha_inicio": str(hoy()), "fecha_fin": str(hoy() - timedelta(days=1))}

    resp = client.get("/reportes/attendance", params=params, headers=_token(admin))

    assert resp.status_code == 422


def test_router_export_formato_invalido_422(db):
    admin = _crear_staff(db, RolUsuario.administrador, "admin-fmt@gymflow.test")
    params = {"fecha_inicio": str(hoy()), "fecha_fin": str(hoy()), "format": "pdf"}

    resp = client.get("/reportes/attendance/export", params=params, headers=_token(admin))

    assert resp.status_code == 422
